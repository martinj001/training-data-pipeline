"""
Ingest manual Excel logs into data/manual.db.

Usage:
  python src/manual/sync.py              # ingest last 30 days
  python src/manual/sync.py --days 90   # ingest last 90 days
  python src/manual/sync.py --days 0    # ingest all time

Re-ingests a file only when its mtime has changed since the last sync —
safe to re-run, and edits to an already-synced file won't be silently
dropped.
"""
import sqlite3
import argparse
from datetime import date, timedelta
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "data" / "manual.db"
STRENGTH_DIR = PROJECT_ROOT / "data" / "manual" / "strength"
BODY_METRICS_FILE = PROJECT_ROOT / "data" / "manual" / "body_metrics.xlsx"


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS strength_sessions (
            date        TEXT NOT NULL,
            exercise    TEXT NOT NULL,
            set_num     INTEGER,
            reps        INTEGER,
            weight      TEXT,
            rpe         REAL,
            done        TEXT,
            time_sec    INTEGER,
            notes       TEXT
        );

        CREATE TABLE IF NOT EXISTS body_metrics (
            date        TEXT PRIMARY KEY,
            weight_kg   REAL,
            calories    INTEGER,
            protein_g   INTEGER,
            sleep_hrs   REAL,
            notes       TEXT
        );

        CREATE TABLE IF NOT EXISTS sync_meta (
            source      TEXT PRIMARY KEY,
            mtime       REAL NOT NULL
        );
    """)
    conn.commit()


def _synced_mtime(conn: sqlite3.Connection, source: str) -> float | None:
    row = conn.execute("SELECT mtime FROM sync_meta WHERE source = ?", (source,)).fetchone()
    return row[0] if row else None


def _mark_synced(conn: sqlite3.Connection, source: str, mtime: float) -> None:
    conn.execute(
        "INSERT INTO sync_meta (source, mtime) VALUES (?, ?) "
        "ON CONFLICT(source) DO UPDATE SET mtime = excluded.mtime",
        (source, mtime),
    )


def ingest_strength(conn: sqlite3.Connection, since: date) -> tuple[int, int]:
    files = sorted(STRENGTH_DIR.glob("*.xlsx"))
    ingested, skipped = 0, 0

    for f in files:
        if f.stem.endswith(".example"):
            continue
        try:
            file_date = date.fromisoformat(f.stem)
        except ValueError:
            continue
        if since and file_date < since:
            continue
        if file_date > date.today():
            continue

        source = f"strength:{f.stem}"
        mtime = f.stat().st_mtime
        if _synced_mtime(conn, source) == mtime:
            skipped += 1
            continue

        conn.execute("DELETE FROM strength_sessions WHERE date = ?", (f.stem,))

        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows_inserted = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            exercise = row[0]
            if not exercise:
                continue
            conn.execute(
                """INSERT INTO strength_sessions
                   (date, exercise, set_num, reps, weight, rpe, done, time_sec, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    f.stem,
                    str(exercise),
                    row[1],
                    row[2],
                    str(row[3]) if row[3] is not None else None,
                    row[4],
                    str(row[5]) if row[5] is not None else None,
                    row[6],
                    str(row[7]) if row[7] is not None else None,
                ),
            )
            rows_inserted += 1
        wb.close()
        _mark_synced(conn, source, mtime)
        conn.commit()
        ingested += 1
        print(f"  strength: {f.stem} ({rows_inserted} rows)")

    return ingested, skipped


def ingest_body_metrics(conn: sqlite3.Connection, since: date) -> tuple[int, int]:
    if not BODY_METRICS_FILE.exists():
        print("  body_metrics.xlsx not found — skipping")
        return 0, 0

    source = "body_metrics"
    mtime = BODY_METRICS_FILE.stat().st_mtime
    if _synced_mtime(conn, source) == mtime:
        skipped = conn.execute(
            "SELECT COUNT(*) FROM body_metrics WHERE date >= ?",
            (since.isoformat(),),
        ).fetchone()[0]
        return 0, skipped

    # File changed since last sync — refresh everything in the window so
    # edits to already-synced rows aren't silently dropped.
    conn.execute("DELETE FROM body_metrics WHERE date >= ?", (since.isoformat(),))

    wb = openpyxl.load_workbook(BODY_METRICS_FILE, read_only=True, data_only=True)
    ws = wb.active
    ingested, skipped = 0, 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row[0]:
            continue
        row_date = str(row[0])[:10]
        try:
            parsed = date.fromisoformat(row_date)
        except ValueError:
            continue
        if since and parsed < since:
            continue

        conn.execute(
            """INSERT INTO body_metrics (date, weight_kg, calories, protein_g, sleep_hrs, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (row_date, row[1], row[2], row[3], row[4], str(row[5]) if row[5] else None),
        )
        ingested += 1
        print(f"  body_metrics: {row_date}")

    wb.close()
    _mark_synced(conn, source, mtime)
    conn.commit()
    return ingested, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest manual Excel logs into manual.db")
    parser.add_argument("--days", type=int, default=30,
                        help="How many days back to ingest (0 = all time, default 30)")
    args = parser.parse_args()

    since = (date.today() - timedelta(days=args.days)) if args.days > 0 else date.min

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    init_db(conn)

    print(f"Manual sync — since {since if args.days > 0 else 'all time'}")

    s_in, s_skip = ingest_strength(conn, since)
    b_in, b_skip = ingest_body_metrics(conn, since)

    conn.close()

    print(f"\nDone — strength: {s_in} new session(s), {s_skip} already present")
    print(f"     — body metrics: {b_in} new day(s), {b_skip} already present")


if __name__ == "__main__":
    main()
