"""
Generate pre-populated strength log workbooks from the current training plan.

Usage:
  python src/manual/build_workbooks.py            # auto-detects most recent plan
  python src/manual/build_workbooks.py --plan data/plans/2026-06-08.md
  python src/manual/build_workbooks.py --force    # overwrite existing files

Creates data/manual/strength/YYYY-MM-DD.xlsx for each strength day in the plan.
Leaves Done and RPE blank — fill those in after each set.
"""

import argparse
import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT         = Path(__file__).resolve().parents[2]
PLANS_DIR    = ROOT / "data" / "plans"
STRENGTH_DIR = ROOT / "data" / "manual" / "strength"

HEADERS = ["Exercise", "Set", "Reps", "Weight (lbs)", "RPE (1-10)", "Done", "Time (seconds)", "Notes"]

COL_WIDTHS = [26, 5, 6, 14, 11, 6, 15, 30]

HEADER_FILL  = PatternFill("solid", fgColor="2E4057")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
ALT_FILL     = PatternFill("solid", fgColor="F2F2F2")

MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

# ── session templates ─────────────────────────────────────────────────────────
# Each row: (exercise, set_num, reps, weight, time_sec, notes)
# weight: int/float for lbs, "Body Weight" for BW, None for bodyweight-timed
# time_sec: seconds for timed exercises (plank, holds), else None

def session_a(week: int) -> list[tuple]:
    pullup_reps = 7 if week == 1 else 8
    ohp_weight  = 50 if week == 1 else 55
    rows = [
        # warm-up
        ("Mobility", None, None, None, None, None),
        # pull-ups / calf raises superset — do calf raises immediately after each set
        ("Pull-Ups",    1, pullup_reps, "Body Weight", None, None),
        ("Calf Raises", 1, 15,          "Body Weight", None, "2-sec up, pause at top, 2-sec down"),
        ("Pull-Ups",    2, pullup_reps, "Body Weight", None, None),
        ("Calf Raises", 2, 15,          "Body Weight", None, None),
        ("Pull-Ups",    3, pullup_reps, "Body Weight", None, None),
        ("Calf Raises", 3, 15,          "Body Weight", None, None),
        ("Pull-Ups",    4, pullup_reps, "Body Weight", None, "drop a rep on set 4 is fine" if week == 2 else None),
        ("Calf Raises", 4, 15,          "Body Weight", None, "add 25 lb DB when bodyweight feels trivial"),
        # dips
        ("Dips", 1, 9, "Body Weight", None, None),
        ("Dips", 2, 9, "Body Weight", None, None),
        ("Dips", 3, 9, "Body Weight", None, "sub: push-up superset (10 standard + 10 incline) if unavailable"),
        # kb row (single arm)
        ("KB Row", 1, 10, 44, None, None),
        ("KB Row", 2, 10, 44, None, None),
        ("KB Row", 3, 10, 44, None, "drop to 35 if form breaks"),
        # barbell overhead press (standing)
        ("Barbell Overhead Press", 1, 10, ohp_weight, None, None),
        ("Barbell Overhead Press", 2, 10, ohp_weight, None, None),
        ("Barbell Overhead Press", 3, 10, ohp_weight, None, "stay at 50 if form breaks on rep 8" if week == 2 else None),
        # core finisher
        ("Dead Bug",                1, None, None,          30, "alternate arm/leg, lower back glued down"),
        ("Dead Bug",                2, None, None,          30, None),
        ("Dead Bug",                3, None, None,          30, None),
        ("Hollow Hold",             1, None, None,          30, None),
        ("Hollow Hold",             2, None, None,          30, None),
        ("Hollow Hold",             3, None, None,          30, None),
        ("Push-Up to Downward Dog", 1, 10,   "Body Weight", None, None),
        ("Push-Up to Downward Dog", 2, 10,   "Body Weight", None, None),
        ("Push-Up to Downward Dog", 3, 10,   "Body Weight", None, None),
    ]
    return rows


def session_b(week: int) -> list[tuple]:
    rdl_weights = [50, 70, 80, 95]  if week == 1 else [55, 75, 85, 100]
    bss_weight  = 35                 if week == 1 else 44   # KB held at chest
    squat_last  = 150                if week == 1 else 150
    squat_note  = "try 150 — 145 felt clean in block 3 (RPE 7.5)" if week == 1 else "hold 150 or try 155 if week 1 felt clean"
    rows = [
        # warm-up
        ("Mobility", None, None, None, None, None),
        # squat
        ("Squat", 1, 10, 95,         None, None),
        ("Squat", 2, 10, 115,        None, None),
        ("Squat", 3, 10, 135,        None, None),
        ("Squat", 4, 10, squat_last, None, squat_note),
        # rdl
        ("Romanian Deadlift", 1, 10, rdl_weights[0], None, None),
        ("Romanian Deadlift", 2, 10, rdl_weights[1], None, None),
        ("Romanian Deadlift", 3, 10, rdl_weights[2], None, None),
        ("Romanian Deadlift", 4, 10, rdl_weights[3], None, "own the hinge"),
        # kb swings
        ("Kettlebell Swing", 1, 15, 35, None, "hip drive, not a squat"),
        ("Kettlebell Swing", 2, 15, 35, None, None),
        ("Kettlebell Swing", 3, 15, 35, None, None),
        # bulgarian split squats / calf raises superset
        ("Bulgarian Split Squat", 1, 8, bss_weight,    None, "each leg — rear foot elevated; find your footing first set"),
        ("Calf Raises",           1, 15, "Body Weight", None, "2-sec up, pause at top, 2-sec down"),
        ("Bulgarian Split Squat", 2, 8, bss_weight,    None, None),
        ("Calf Raises",           2, 15, "Body Weight", None, None),
        ("Bulgarian Split Squat", 3, 8, bss_weight,    None, None),
        ("Calf Raises",           3, 15, "Body Weight", None, None),
        # core finisher
        ("Plank",               1, None, None,          90, "posterior tilt — tuck pelvis"),
        ("Plank",               2, None, None,          90, None),
        ("Plank",               3, None, None,          90, None),
        ("Hanging Knee Raises", 1, 10,   "Body Weight", None, "on pull-up bar — progress to straight-leg when easy"),
        ("Hanging Knee Raises", 2, 10,   "Body Weight", None, None),
        ("Hanging Knee Raises", 3, 10,   "Body Weight", None, None),
    ]
    return rows


SESSION_BUILDERS = {"A": session_a, "B": session_b}


# ── plan parsing ──────────────────────────────────────────────────────────────

def parse_strength_days(plan_path: Path) -> list[tuple[date, str]]:
    """Return [(session_date, 'A' or 'B'), ...] for all strength rows in the plan."""
    plan_year = int(plan_path.stem[:4])
    results   = []

    for line in plan_path.read_text(encoding="utf-8").splitlines():
        if "| Strength |" not in line and "| strength |" not in line.lower():
            continue
        # Expect table row: | June 8 | Mon | Strength | Session A ... | ... |
        parts = [p.strip() for p in line.split("|") if p.strip()]
        if len(parts) < 4:
            continue
        date_str    = parts[0]   # e.g. "June 8"
        session_str = parts[3]   # e.g. "Session A — Upper / Push-Pull"

        session_match = re.search(r"Session ([AB])", session_str, re.IGNORECASE)
        if not session_match:
            continue

        date_match = re.match(r"(\w+)\s+(\d+)", date_str)
        if not date_match:
            continue

        month_name = date_match.group(1).lower()
        day        = int(date_match.group(2))
        month      = MONTH_MAP.get(month_name)
        if not month:
            continue

        try:
            session_date = date(plan_year, month, day)
        except ValueError:
            continue

        results.append((session_date, session_match.group(1).upper()))

    return results


# ── workbook builder ──────────────────────────────────────────────────────────

def build_workbook(out_path: Path, rows: list[tuple], session_label: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = session_label

    # header
    ws.append(HEADERS)
    for col_idx, (cell, width) in enumerate(zip(ws[1], COL_WIDTHS), start=1):
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.freeze_panes = "A2"

    # data rows
    prev_exercise = None
    use_alt       = False

    for exercise, set_num, reps, weight, time_sec, notes in rows:
        if exercise != prev_exercise:
            use_alt      = not use_alt
            prev_exercise = exercise

        ws.append([exercise, set_num, reps, weight, None, None, time_sec, notes])

        row_num = ws.max_row
        fill    = ALT_FILL if use_alt else None
        for cell in ws[row_num]:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left")

    # center numeric columns
    for row in ws.iter_rows(min_row=2):
        for cell in row[1:6]:  # Set, Reps, Weight, RPE, Done
            cell.alignment = Alignment(horizontal="center")

    wb.save(out_path)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Build pre-populated strength workbooks")
    parser.add_argument("--plan",  help="Path to plan markdown file (default: most recent)")
    parser.add_argument("--force", action="store_true", help="Overwrite existing files")
    args = parser.parse_args()

    if args.plan:
        plan_path = Path(args.plan)
    else:
        plans = sorted(PLANS_DIR.glob("*.md"))
        if not plans:
            raise SystemExit("No plan files found in data/plans/")
        plan_path = plans[-1]

    print(f"Plan: {plan_path.name}")
    STRENGTH_DIR.mkdir(parents=True, exist_ok=True)

    strength_days = parse_strength_days(plan_path)
    if not strength_days:
        raise SystemExit("No strength sessions found in plan — check table format.")

    plan_start = date.fromisoformat(plan_path.stem)

    for session_date, session_type in strength_days:
        out_path = STRENGTH_DIR / f"{session_date.isoformat()}.xlsx"
        if out_path.exists() and not args.force:
            print(f"  skip  {out_path.name}  (already exists — use --force to overwrite)")
            continue

        # week 1 = days 0-6, week 2 = days 7+
        delta = (session_date - plan_start).days
        week  = 1 if delta < 7 else 2

        builder = SESSION_BUILDERS[session_type]
        rows    = builder(week)
        label   = f"Session {session_type} — Week {week}"
        build_workbook(out_path, rows, label)
        print(f"  built {out_path.name}  ({label})")

    print("Done.")


if __name__ == "__main__":
    main()
